#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
build_macola_and_fg_bom_seed.py
===============================
Emits sql/scratch/seed_macola_and_fg_boms.sql (plus a row-level reconciliation
CSV) from the customer workbook

    reference/MPP_Macola_Numbers_2026-06-15.xlsx
    (delivered as "MACOLA NUMBERS FOR INVENTORYupdate 6-15-26.xlsx")

The workbook is the single source of truth for Macola numbers: edit it (or drop
in MPP's next revision) and re-run this, never hand-edit the emitted SQL.

THE GOVERNING RULE
------------------
A Macola number is ONLY a value from a column literally headed "MACOLA #".
Exactly two sheets have one: ALUMINUM (C1) and SUPPLY PARTS (C1).

The per-family sheets (59B 6MA CH / OP / RPY 5BA 6B2 / FUEL PUMP 5G0 / 5A2 /
SERVICE / PASS THRU / NEW MODEL) carry columns headed RAW / TUMBLED BLASTED /
MACHINED / FINISHED GOODS holding values like 187-090, 187-091, 187-092,
187-MET, 186-AEP. Those are MPP per-stage inventory codes. They are NOT Macola
numbers and this script NEVER writes them to Parts.Item.MacolaPartNumber. The
family sheets are read for ONE purpose only: their PART NAME -> CUSTOMER PART #
pairs, used as a lookup table to turn the free-text FG names on SUPPLY PARTS
into real Honda part numbers.

WHAT IT EMITS
-------------
  1. Parts.Item      -- the 4 ALUMINUM alloys as ItemType RawMaterial, emitted
                        COMMENTED OUT. Raw Material Tracking is FUTURE scope
                        (MPP_Scope_Matrix.xlsx row 21 / FRS 3.9.1), and FUTURE
                        means "do not populate". Their Macola numbers therefore
                        do not load. Un-comment when scope changes.
  2. Parts.Item      -- UPDATE MacolaPartNumber, and ONLY where it is currently
                        NULL or blank. A non-blank value is never overwritten.
                        No Component/FinishedGood rows are ever CREATED from
                        this workbook (it is an inventory list, not a part
                        master; inventing rows would mint near-duplicate part
                        numbers such as 92900-06012-1B vs 92900-06012-0B).
  3. Parts.Bom /
     Parts.BomLine   -- finished-good BOM lines derived from SUPPLY PARTS
                        (component + "Pcs per part" -> "Corresponding FG
                        Assembly(s)"), published v1, for the pairs where BOTH
                        sides resolve to a catalog part number.

Matching baseline is reference/MPP_Seed_Layout_Proposal.xlsx (Items + BOMs
tabs) -- the same reviewed workbook build_mpp_parts_seed.py emits from -- so
this script needs no database connection and stays reproducible.

Emitted SQL follows sql/seeds/020_seed_items.sql + sql/scratch/seed_mpp_parts.sql:
  * idempotent - IF NOT EXISTS / NOT EXISTS on natural keys (PartNumber),
    never a hardcoded Id, so it is safe to re-run in place.
  * ASCII-only string literals (sqlcmd reads .sql in the Windows codepage, so a
    non-ASCII byte lands in the DB as mojibake). Asserted below, not assumed --
    the source workbook definitely contains non-ASCII (embedded newlines at
    minimum).
  * set-based via table variables, so no per-entity GO/DECLARE churn.

Usage:  python reference/scripts/build_macola_and_fg_bom_seed.py [workbook.xlsx]
"""
import csv
import os
import re
import sys
import warnings
from collections import defaultdict

warnings.filterwarnings('ignore')
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SRC = os.path.join(ROOT, 'reference', 'MPP_Macola_Numbers_2026-06-15.xlsx')
CATALOG = os.path.join(ROOT, 'reference', 'MPP_Seed_Layout_Proposal.xlsx')
OUT = os.path.join(ROOT, 'sql', 'scratch', 'seed_macola_and_fg_boms.sql')
CSV_OUT = os.path.join(ROOT, 'sql', 'scratch', 'macola_bom_reconciliation.csv')
if len(sys.argv) > 1:
    SRC = sys.argv[1]

# Sheets that carry a real "MACOLA #" header.
MACOLA_SHEETS = ('ALUMINUM', 'SUPPLY PARTS')
# Per-family sheets, read ONLY for PART NAME -> CUSTOMER PART #.
FAMILY_SHEETS = ('59B 6MA CH', 'OP', 'RPY 5BA 6B2 ', 'FUEL PUMP 5G0 ', '5A2',
                 'SERVICE', 'PASS THRU', 'NEW MODEL')

# SCOPE GATE. reference/MPP_Scope_Matrix.xlsx row 21 -- Traceability / Raw
# Material Tracking -- is "Not Included", Future (MPP_MES_SUMMARY.md line 307:
# "FUTURE, excluded per FRS 3.9.1"). CLAUDE.md's FUTURE rule is "schema supports
# it, but do NOT implement, POPULATE, or test". So the ALUMINUM sheet's four
# alloys are NOT created as Parts.Item rows: their SQL is emitted COMMENTED OUT,
# ready to un-comment the day raw-material tracking comes into scope.
# LB because MPP charges melt by weight -- confirm with MPP before enabling.
ALLOY_ITEM_TYPE = 'RawMaterial'
ALLOY_UOM = 'LB'
ALLOY_SEED_ENABLED = False


# ------------------------------------------------------------------ helpers --
def sheet_rows(ws):
    """Every row as a list of stripped strings; embedded newlines -> spaces."""
    out = []
    for row in ws.iter_rows(values_only=True):
        out.append(['' if c is None else re.sub(r'\s+', ' ', str(c)).strip() for c in row])
    return out


def cell(row, i):
    return row[i] if i is not None and i < len(row) else ''


MODEL_CODE = re.compile(r'^[0-9][A-Z0-9]{2}$')


def name_key(s):
    """Normalize a free-text part NAME for lookup.

    Typography only -- no judgement:
      * upper-case, every non-alphanumeric run -> a token break
      * a 3-char token starting with a digit is a Honda model code, so letter O
        folds to zero (5GO -> 5G0). Word tokens (OIL, ROCKER) are untouched.
      * trailing plural S stripped (OIL PANS == OIL PAN, CH/RS SETS == CH/RS SET)
      * tokens joined with no separator, so a missing space is absorbed
        (P8A OILPAN == P8A OIL PAN, RNAOIL PAN == RNA OIL PAN)
    """
    s = re.sub(r'[^A-Z0-9]+', ' ', (s or '').upper()).strip()
    toks = [t.replace('O', '0') if MODEL_CODE.match(t) else t for t in s.split()]
    k = ''.join(toks)
    while k.endswith('S') and len(k) > 3:
        k = k[:-1]
    return k


def pn_key(s):
    """Normalize a part NUMBER for lookup.

    Strip every separator (the catalog pads Honda numbers to fixed width, so
    '1120A-64A -A000' and '1120A64AA000' are the same number), then fold the
    letter O in the third slot of a Honda model code -- <digit><letter>O -- to
    zero. The catalog writes the 5G0 family as 5GO and the 5R0 dowel as 5RO,
    where MPP's inventory workbook writes 5G0 / 5R0. Narrow on purpose: an O
    anywhere else (030ES-M4464, 90703-PEO-0000) is left alone.
    """
    k = re.sub(r'[^A-Z0-9]', '', (s or '').upper())
    return re.sub(r'(?<=[0-9][A-Z])O', '0', k)


def q(s):
    """Single-quoted NVARCHAR literal, ASCII-verified."""
    s = '' if s is None else str(s)
    try:
        s.encode('ascii')
    except UnicodeEncodeError:
        BAD.append(s)
    return "N'" + s.replace("'", "''") + "'"


def num(s):
    s = ('' if s is None else str(s)).strip()
    try:
        f = float(s)
    except ValueError:
        return None
    return int(f) if f == int(f) else f


BAD = []

# ------------------------------------------------------------------- inputs --
wb = openpyxl.load_workbook(SRC, data_only=True)
cat_wb = openpyxl.load_workbook(CATALOG, data_only=True)

# ---- catalog: the reviewed part master (no DB needed) ----
catalog = []
for r in cat_wb['Items'].iter_rows(min_row=2, values_only=True):
    if r[0]:
        catalog.append((str(r[0]).strip(), str(r[1] or '').strip(), str(r[2] or '').strip()))
# The 13-part demo matrix from sql/seeds/020_seed_items.sql (plus the AIM
# post-back test part) lives alongside the catalog in every dev/prod DB, so it
# counts as a match target too. Listed in FULL on purpose: several demo part
# numbers are string-distinct near-twins of a catalog row -- 90701-5R0-3000
# (demo, zero) vs 90701-5RO-3000 (catalog, letter O) are two separate rows for
# the same dowel pin -- and including both makes the match AMBIGUOUS, so the
# generator reports it instead of picking one.
DEMO = [
    ('5G0-c', '5G0 Front Cover Casting', 'Component'),
    ('5G0-SA', '5G0 Front Cover Sub-Assembly', 'SubAssembly'),
    ('5G0-FG', '5G0 Front Cover Finished Good', 'FinishedGood'),
    ('21001 pin', 'Pin 21001', 'Component'),
    ('12231-59B-0000', '59B Cam Holder IN #1 Casting', 'Component'),
    ('12232-59B-0000', '59B Cam Holder IN #2 Casting', 'Component'),
    ('12241-59B-0000', '59B Cam Holder EX #1 Casting', 'Component'),
    ('90701-5R0-3000', 'Dowel Pin 9x10 (purchased)', 'Component'),
    ('1223A-59B -A0002', '59B Cam-Rocker Holder Set', 'FinishedGood'),
    ('12270-6NA', '6NA Fuel Pump Base Casting (raw)', 'Component'),
    ('12270-6NA-M', '6NA Fuel Pump Base Machined (synth SA)', 'SubAssembly'),
    ('92900-06014-1B', 'Stud Bolt 6x14 (purchased)', 'Component'),
    ('94301-08100', 'Dowel Pin 8x10 (purchased)', 'Component'),
    ('12270-6NA -0001', '6NA Fuel Pump', 'FinishedGood'),
    ('11200-6FB -A000', '6FB Front Cover (AIM post-back test part)', 'FinishedGood'),
]
known = {c[0] for c in catalog}
catalog += [d for d in DEMO if d[0] not in known]

by_pn = defaultdict(list)      # pn_key   -> [(PartNumber, Description, ItemType)]
by_desc = defaultdict(set)     # name_key -> {PartNumber}
for pn, desc, itype in catalog:
    by_pn[pn_key(pn)].append((pn, desc, itype))
    if desc:
        by_desc[name_key(desc)].add(pn)

# A normalization that collapses two DIFFERENT catalog parts onto one key would
# silently mis-assign a Macola number. Surface it rather than trust it.
PN_COLLISIONS = {k: [h[0] for h in v] for k, v in by_pn.items() if len(v) > 1}

# ---- catalog BOM content, so we can classify AGREE / NEW / QTY-CONFLICT ----
existing_bom = {}   # (parent, child) -> qty
for r in cat_wb['BOMs'].iter_rows(min_row=2, values_only=True):
    if r[0] and r[2]:
        existing_bom[(str(r[0]).strip(), str(r[2]).strip())] = str(r[4] or '').strip()
bom_parents = {p for p, _ in existing_bom}

# ---- ALUMINUM: Part Description | Honda Part# | MACOLA # | Corresponding FG ----
al_rows = sheet_rows(wb['ALUMINUM'])
al_hdr = [c.upper() for c in al_rows[0]]
assert 'MACOLA #' in al_hdr, "ALUMINUM has no 'MACOLA #' header -- refusing to guess"
AL_MAC = al_hdr.index('MACOLA #')
AL_PN = al_hdr.index('HONDA PART#')
AL_DESC = al_hdr.index('PART DESCRIPTION')
AL_FG = al_hdr.index('CORRESPONDING FG ASSEMBLY(S)')
alloys = []
for i, r in enumerate(al_rows[1:], start=2):
    if cell(r, AL_MAC):
        alloys.append({'row': i, 'desc': cell(r, AL_DESC), 'pn': cell(r, AL_PN),
                       'macola': cell(r, AL_MAC), 'fg': cell(r, AL_FG)})

# ---- SUPPLY PARTS: repeated header bands + blank-continuation forward fill ----
sp_rows = sheet_rows(wb['SUPPLY PARTS'])
sp_hdr = [c.upper() for c in sp_rows[0]]
assert 'MACOLA #' in sp_hdr, "SUPPLY PARTS has no 'MACOLA #' header -- refusing to guess"
SP_DESC, SP_PN = sp_hdr.index('SUPPLY PART DESCRIPTION'), sp_hdr.index('HONDA PART#')
SP_MAC, SP_FG = sp_hdr.index('MACOLA #'), sp_hdr.index('CORRESPONDING FG ASSEMBLY(S)')
SP_QTY = sp_hdr.index('PCS PER PART')
HDR_MARK = sp_hdr[SP_DESC]

supply = []   # one dict per anchor row (a row that carries a MACOLA #)
links = []    # one dict per component -> FG link (anchor + forward-filled rows)
header_rows = []
anchor = None
for i, r in enumerate(sp_rows, start=1):
    if cell(r, SP_DESC).upper() == HDR_MARK:
        header_rows.append(i)                       # a repeated header band
        continue
    if not any(r):
        continue
    if cell(r, SP_MAC):                             # anchor row
        anchor = {'row': i, 'desc': cell(r, SP_DESC), 'pn': cell(r, SP_PN),
                  'macola': cell(r, SP_MAC), 'qty': cell(r, SP_QTY)}
        supply.append(anchor)
    if anchor is None:
        continue
    fg = cell(r, SP_FG)
    if fg:
        # Continuation rows leave the first three columns blank; the qty column
        # is blank on most of them but carries a per-FG override on 20 of them.
        # Blank qty therefore FORWARD-FILLS from the anchor.
        links.append({'row': i, 'anchor': anchor, 'fg': fg,
                      'qty': cell(r, SP_QTY) or anchor['qty'],
                      'qty_own': bool(cell(r, SP_QTY)),
                      'continuation': not cell(r, SP_MAC)})

# ---- family sheets: PART NAME -> CUSTOMER PART # (block-wise, many headers) --
fam_pairs = []          # (sheet, row, part name, customer pn)
fam_blocks = []         # (sheet, header row, block label) -- structure report
by_fam_name = defaultdict(set)
for sname in FAMILY_SHEETS:
    hdr = None
    for i, r in enumerate(sheet_rows(wb[sname]), start=1):
        up = [c.upper() for c in r]
        if 'RAW' in up and 'FINISHED GOODS' in up:      # a new header band
            hdr = {c: j for j, c in enumerate(up) if c}
            fam_blocks.append((sname, i, cell(r, 0)))
            continue
        if hdr is None:
            continue

        def g(*names, _r=r, _h=hdr):
            for n in names:
                if n in _h and _h[n] < len(_r):
                    return _r[_h[n]]
            return ''
        nm = g('PART NAME')
        cp = g('CUSTOMER PART #', 'CUSTOMER                 PART #')
        if nm:
            fam_pairs.append((sname, i, nm, cp))
            if cp:
                by_fam_name[name_key(nm)].add(cp)


def resolve_item(part_number):
    """Honda part number -> catalog PartNumber. Unique hit or nothing."""
    hits = by_pn.get(pn_key(part_number), []) if part_number else []
    return sorted({h[0] for h in hits})


def resolve_fg(fg_name):
    """Free-text FG assembly name -> catalog PartNumber, via two ladders.

    (a) catalog Description, (b) family-sheet PART NAME -> CUSTOMER PART # ->
    catalog PartNumber. Returns (resolved, via, dangling) -- 'dangling' lists
    family-sheet customer part numbers that are absent from the catalog, which
    is the single biggest reason a name fails to resolve.
    """
    k = name_key(fg_name)
    resolved, via, dangling = set(), [], []
    if k in by_desc:
        resolved |= by_desc[k]
        via.append('catalog-description')
    if k in by_fam_name:
        via.append('family-sheet-name')
        for cp in by_fam_name[k]:
            hits = resolve_item(cp)
            if hits:
                resolved |= set(hits)
            else:
                dangling.append(cp)
    return sorted(resolved), '+'.join(via), sorted(set(dangling))


# --------------------------------------------------------- 1. Macola numbers --
macola_updates = []     # (PartNumber, Macola, source sheet, source row, note)
macola_unmatched = []   # rows whose Honda Part# has no catalog counterpart

# The 4 ALUMINUM alloys are NOT counted as landed -- raw-material tracking is
# FUTURE scope, so their SQL is emitted commented out (see above).
for s in supply:
    hits = resolve_item(s['pn'])
    if len(hits) == 1:
        raw = re.sub(r'[^A-Z0-9]', '', s['pn'].upper())
        note = '' if re.sub(r'[^A-Z0-9]', '', hits[0].upper()) == raw else 'O/0 fold'
        macola_updates.append((hits[0], s['macola'], 'SUPPLY PARTS', s['row'], note))
    else:
        macola_unmatched.append(s)

# ------------------------------------------------------------ 2. FG BOM lines --
bom_ok, bom_conflict = [], []
bom_parent_miss, bom_child_miss, bom_both_miss = [], [], []
for ln in links:
    a = ln['anchor']
    child = resolve_item(a['pn'])
    parent, via, dangling = resolve_fg(ln['fg'])
    qty = num(ln['qty'])
    if len(child) == 1 and len(parent) == 1 and qty is not None:
        key = (parent[0], child[0])
        prior = existing_bom.get(key)
        rec = {'row': ln['row'], 'parent': parent[0], 'child': child[0], 'qty': qty,
               'fg': ln['fg'], 'desc': a['desc'], 'via': via, 'prior': prior}
        if prior is not None and num(prior) != qty:
            bom_conflict.append(rec)
        else:
            rec['state'] = 'agrees-with-catalog' if prior is not None else 'new'
            bom_ok.append(rec)
    elif len(child) == 1:
        bom_parent_miss.append({'row': ln['row'], 'fg': ln['fg'], 'anchor': a,
                                'qty': ln['qty'], 'dangling': dangling, 'via': via})
    elif len(parent) == 1:
        bom_child_miss.append({'row': ln['row'], 'fg': ln['fg'], 'anchor': a, 'qty': ln['qty']})
    else:
        bom_both_miss.append({'row': ln['row'], 'fg': ln['fg'], 'anchor': a,
                              'qty': ln['qty'], 'dangling': dangling})

# BOM lines actually worth emitting: the new ones (an agreeing line is already
# in the catalog seed, and re-emitting it is harmless but noisy).
bom_emit = [b for b in bom_ok if b['state'] == 'new']
bom_emit.sort(key=lambda b: (b['parent'], b['child']))
sort_seq = defaultdict(int)
for b in bom_emit:
    sort_seq[b['parent']] += 1
    b['sort'] = sort_seq[b['parent']]

# ---------------------------------------------------------- reverse direction --
got_macola = {u[0] for u in macola_updates}
prior_macola = {}
for r in cat_wb['Items'].iter_rows(min_row=2, values_only=True):
    if r[0] and r[11]:
        prior_macola[str(r[0]).strip()] = str(r[11]).strip()
no_macola = [c for c in catalog if c[0] not in got_macola and c[0] not in prior_macola]
fg_total = [c for c in catalog if c[2] == 'FinishedGood']
fg_with = [c for c in fg_total if c[0] in got_macola]

# ==================================================================== emit SQL --
L = []
A = L.append

A("""-- ============================================================
-- Seed:        seed_macola_and_fg_boms.sql   (GENERATED -- edit
--              reference/scripts/build_macola_and_fg_bom_seed.py, never this file)
-- Author:      Blue Ridge Automation
-- Date:        2026-08-18
-- Description: Macola part numbers + finished-good BOM lines, transformed from
--              the customer workbook reference/MPP_Macola_Numbers_2026-06-15.xlsx
--              ("MACOLA NUMBERS FOR INVENTORYupdate 6-15-26.xlsx", MPP 6/15/26).
--
--              Lives in sql/scratch/ (NOT sql/seeds/) for the same reason
--              seed_mpp_parts.sql does: it is a first load that still carries
--              open customer data questions (see the RECONCILIATION appendix at
--              the bottom of this file and the report that generated it), and
--              dropping it into every Reset-DevDatabase / test-DB build before
--              MPP answers them would destabilise the suite. Promote to
--              sql/seeds/ once the data is settled.
--
--              THE GOVERNING RULE. A Macola number comes ONLY from a column
--              literally headed "MACOLA #". Exactly two sheets have one:
--              ALUMINUM (C1) and SUPPLY PARTS (C1). The per-family sheets'
--              RAW / TUMBLED BLASTED / MACHINED / FINISHED GOODS columns
--              (187-090, 187-091, 187-092, 187-MET, 186-AEP ...) are MPP
--              per-stage inventory codes, NOT Macola numbers, and nothing in
--              this file writes one to Parts.Item.MacolaPartNumber.
--
--              SCOPE. Nothing here CREATES a Parts.Item row. The workbook is an
--              inventory list, not a part master, and minting rows from it would
--              produce near-duplicate part numbers (92900-06012-1B alongside the
--              catalog's 92900-06012-0B). The four ALUMINUM alloys are the one
--              case where a Macola number has no item to attach to; their SQL is
--              emitted COMMENTED OUT because Raw Material Tracking is FUTURE
--              (MPP_Scope_Matrix.xlsx row 21 / FRS 3.9.1). The SERVICE /
--              PASS THRU / NEW MODEL sheets carry no MACOLA # column and are read
--              only as a PART NAME -> CUSTOMER PART # lookup table.
--
--              Fully idempotent -- re-running inserts and updates nothing twice.
--              An Item that ALREADY carries a non-blank MacolaPartNumber is
--              LEFT ALONE, never overwritten. ASCII-only.
--
--              Dependencies: 0005 (Parts.Item.MacolaPartNumber + its filtered
--              index), 0007 (Parts.Bom / Parts.BomLine), 004 (ItemType/Uom),
--              and a loaded part catalog (sql/scratch/seed_mpp_parts.sql and/or
--              sql/seeds/020_seed_items.sql). Rows whose PartNumber is absent
--              simply update/insert nothing -- the seed does not create them.
-- ============================================================
SET NOCOUNT ON;
SET XACT_ABORT ON;
GO
""")

# ---- 1. aluminum alloys -- OUT OF SCOPE, emitted commented out ----
alloy_sql = [
    "DECLARE @Dev BIGINT = (SELECT TOP 1 Id FROM Location.AppUser WHERE Initials = N'DEV' ORDER BY Id);",
    "DECLARE @A TABLE (Pn NVARCHAR(50), Descr NVARCHAR(500), Macola NVARCHAR(50));",
    "INSERT INTO @A (Pn, Descr, Macola) VALUES",
]
alloy_sql += [
    " (%s, %s, %s)%s" % (q(a['pn']), q(a['desc']), q(a['macola']),
                         ';' if i == len(alloys) - 1 else ',')
    for i, a in enumerate(alloys)
]
alloy_sql += [
    "",
    "INSERT INTO Parts.Item (ItemTypeId, PartNumber, Description, MacolaPartNumber,",
    "                        UomId, CreatedAt, CreatedByUserId)",
    "SELECT t.Id, a.Pn, a.Descr, a.Macola, u.Id, SYSUTCDATETIME(), @Dev",
    "FROM @A a",
    "JOIN Parts.ItemType t ON t.Code = %s" % q(ALLOY_ITEM_TYPE),
    "JOIN Parts.Uom      u ON u.Code = %s" % q(ALLOY_UOM),
    "WHERE NOT EXISTS (SELECT 1 FROM Parts.Item x WHERE x.PartNumber = a.Pn);",
    "PRINT 'seed_macola: RawMaterial alloys -> ' + CAST(@@ROWCOUNT AS VARCHAR(10)) + ' inserted.';",
    "GO",
]
A("""-- ============================================================
-- ALUMINUM sheet -- the 4 alloys and their Macola numbers (300 / 301 / 303 /
-- 304). NOT LOADED. The SQL below is deliberately COMMENTED OUT.
--
-- WHY. The ALUMINUM sheet's "Honda Part#" column holds alloy codes -- ADC12,
-- HD2BS, HD2G, NH41 -- not Honda part numbers, and nothing in the catalog
-- corresponds to them. Their Macola numbers have nowhere to land unless the
-- alloys exist as Parts.Item rows of ItemType 'RawMaterial'. But
-- reference/MPP_Scope_Matrix.xlsx row 21 -- Traceability / Raw Material
-- Tracking -- is "Not Included", Future (MPP_MES_SUMMARY.md line 307: FUTURE,
-- excluded per FRS 3.9.1), and the FUTURE rule is "schema supports it, but do
-- NOT implement, POPULATE, or test". Parts.ItemType 'RawMaterial' (Id 1,
-- migration 0004) therefore stays empty.
--
-- To enable, once raw-material tracking is in scope: un-comment the block
-- below and confirm the UoM (LB assumed -- MPP charges melt by weight).
-- ============================================================""")
A('\n'.join(('-- ' + ln).rstrip() for ln in alloy_sql))
A('')
assert not ALLOY_SEED_ENABLED, "ALLOY_SEED_ENABLED=True needs the emit path un-commented too"

# ---- 2. Macola numbers ----
supply_updates = [u for u in macola_updates if u[2] == 'SUPPLY PARTS']
A("""-- ============================================================
-- Parts.Item.MacolaPartNumber -- from the SUPPLY PARTS "MACOLA #" column only.
--
-- Written ONLY where the column is currently NULL or blank. A part that already
-- carries a value is skipped, per the catalog seed's "already present is left
-- alone" convention. (Note: %d catalog rows already carry a value that came from
-- the per-family sheets' FINISHED GOODS column -- 186-AEP, 142-AEP, 141-HCM,
-- 630-AEP, 662-AEP -- which under the governing rule are NOT Macola numbers.
-- They are deliberately NOT touched here; correcting them is Hunter's call.)
--
-- PartNumber is the emitted natural key: the generator has already reconciled
-- the workbook's Honda Part# against the catalog's fixed-width rendering
-- (e.g. sheet '14631-5G0-A000' -> catalog '146315GO A000'), so the SQL needs no
-- fuzzy matching of its own. A PartNumber absent from Parts.Item updates
-- nothing -- by design.
-- ============================================================
DECLARE @M TABLE (Pn NVARCHAR(50), Macola NVARCHAR(50));
INSERT INTO @M (Pn, Macola) VALUES""" % len(prior_macola))
A(',\n'.join(" (%s, %s)" % (q(pn), q(mac)) for pn, mac, _s, _r, _n in
             sorted(supply_updates, key=lambda x: x[0])) + ';')
A("""
DECLARE @DevU BIGINT = (SELECT TOP 1 Id FROM Location.AppUser WHERE Initials = N'DEV' ORDER BY Id);
UPDATE i
   SET i.MacolaPartNumber = m.Macola,
       i.UpdatedAt        = SYSUTCDATETIME(),
       i.UpdatedByUserId  = @DevU
FROM Parts.Item i
JOIN @M m ON m.Pn = i.PartNumber
WHERE NULLIF(LTRIM(RTRIM(i.MacolaPartNumber)), N'') IS NULL;
PRINT 'seed_macola: MacolaPartNumber   -> ' + CAST(@@ROWCOUNT AS VARCHAR(10)) + ' updated.';

DECLARE @Absent INT = (SELECT COUNT(*) FROM @M m
                       WHERE NOT EXISTS (SELECT 1 FROM Parts.Item i WHERE i.PartNumber = m.Pn));
IF @Absent > 0
    PRINT 'seed_macola: WARNING -- ' + CAST(@Absent AS VARCHAR(10)) +
          ' part number(s) in this seed are absent from Parts.Item (catalog not loaded?).';
GO
""")

# ---- 3. FG BOMs ----
A("""-- ============================================================
-- Parts.Bom + Parts.BomLine -- finished-good BOMs from SUPPLY PARTS.
--
-- Source shape: an anchor row carries Supply Part Description | Honda Part# |
-- MACOLA # | Corresponding FG Assembly(s) | Pcs per part. A component feeding
-- several finished goods gets CONTINUATION rows underneath with the first three
-- columns blank; those are forward-filled. "Pcs per part" is blank on most
-- continuations (qty forward-fills from the anchor) but carries a per-FG
-- override on 20 of them, which is honoured.
--
-- Emitted only where BOTH sides resolve to a catalog part number. The BOM is
-- Published v1, mirroring 020_seed_items.sql / seed_mpp_parts.sql. A parent
-- that already has a v1 BOM gets these lines ADDED to it; a line already
-- present (same parent + child) is skipped, so nothing is overwritten.
--
-- Lines where the workbook's qty DISAGREES with the qty already in the catalog
-- are deliberately NOT emitted -- they are listed in the RECONCILIATION
-- appendix at the bottom of this file for MPP to arbitrate.
-- ============================================================
DECLARE @Dev BIGINT = (SELECT TOP 1 Id FROM Location.AppUser WHERE Initials = N'DEV' ORDER BY Id);
DECLARE @B TABLE (Parent NVARCHAR(50), Child NVARCHAR(50), Qty DECIMAL(10,4), Sort INT);
INSERT INTO @B (Parent, Child, Qty, Sort) VALUES""")
_bom_vals = [" (%s, %s, %s, %d)" % (q(b['parent']), q(b['child']), b['qty'], b['sort'])
             for b in bom_emit]
A('\n'.join('%s%s  -- %s <- %s' % (v, ';' if i == len(_bom_vals) - 1 else ',',
                                   b['fg'], b['desc'])
            for i, (v, b) in enumerate(zip(_bom_vals, bom_emit))))
A("""
INSERT INTO Parts.Bom (ParentItemId, VersionNumber, EffectiveFrom, PublishedAt, CreatedByUserId, CreatedAt)
SELECT DISTINCT i.Id, 1, '2026-01-15', '2026-01-14', @Dev, SYSUTCDATETIME()
FROM @B b
JOIN Parts.Item i ON i.PartNumber = b.Parent
WHERE NOT EXISTS (SELECT 1 FROM Parts.Bom x WHERE x.ParentItemId = i.Id AND x.VersionNumber = 1);
PRINT 'seed_macola: Parts.Bom          -> ' + CAST(@@ROWCOUNT AS VARCHAR(10)) + ' inserted.';

INSERT INTO Parts.BomLine (BomId, ChildItemId, QtyPer, UomId, SortOrder)
SELECT bm.Id, c.Id, b.Qty, c.UomId,
       (SELECT ISNULL(MAX(x.SortOrder), 0) FROM Parts.BomLine x WHERE x.BomId = bm.Id) + b.Sort
FROM @B b
JOIN Parts.Item p  ON p.PartNumber   = b.Parent
JOIN Parts.Bom  bm ON bm.ParentItemId = p.Id AND bm.VersionNumber = 1
JOIN Parts.Item c  ON c.PartNumber   = b.Child
WHERE NOT EXISTS (SELECT 1 FROM Parts.BomLine x WHERE x.BomId = bm.Id AND x.ChildItemId = c.Id);
PRINT 'seed_macola: Parts.BomLine      -> ' + CAST(@@ROWCOUNT AS VARCHAR(10)) + ' inserted.';
GO
""")

# ---- 4. reconciliation appendix (comments only) ----
def block(title, lines):
    A('-- ' + '-' * 74)
    A('-- ' + title)
    A('-- ' + '-' * 74)
    if not lines:
        A('--   (none)')
    for x in lines:
        A('--   ' + x)
    A('--')


A('/* ============================================================')
A('   RECONCILIATION APPENDIX -- generated with this file. Nothing below')
A('   executes; it is the audit trail for what the workbook could NOT be')
A('   turned into SQL, and why. Row-level detail: macola_bom_reconciliation.csv')
A('   ============================================================')
A('--')
block('COUNTS', [
    'workbook MACOLA # values ............... %d (ALUMINUM %d + SUPPLY PARTS %d)'
    % (len(alloys) + len(supply), len(alloys), len(supply)),
    'catalog items (match baseline) ......... %d' % len(catalog),
    'Macola numbers landed on an item ....... %d' % len(macola_updates),
    'Macola numbers held back (FUTURE scope)  %d   (ALUMINUM -- raw material)' % len(alloys),
    'Macola numbers with no catalog item .... %d' % len(macola_unmatched),
    'catalog items left with no Macola ...... %d' % len(no_macola),
    'finished goods in catalog .............. %d' % len(fg_total),
    'finished goods GIVEN a Macola number ... %d   <-- see the note below'
    % len(fg_with),
    'SUPPLY PARTS component->FG links ....... %d' % len(links),
    'BOM lines resolved end to end .......... %d' % len(bom_ok),
    '  of which new (emitted) ............... %d' % len(bom_emit),
    '  of which already in the catalog ...... %d' % (len(bom_ok) - len(bom_emit)),
    'BOM lines with a QTY CONFLICT .......... %d' % len(bom_conflict),
    'BOM links, FG parent unresolved ........ %d' % len(bom_parent_miss),
    'BOM links, component child unresolved .. %d' % len(bom_child_miss),
    'BOM links, neither side resolved ....... %d' % len(bom_both_miss),
])
block('THE FINISHED-GOOD / MACOLA TENSION', [
    'The task was phrased "add the Macola numbers to finished goods", but the',
    'only MACOLA # columns in this workbook are on ALUMINUM (raw material) and',
    'SUPPLY PARTS (purchased components). Neither is a finished good. Under the',
    'governing rule, finished goods therefore receive NO Macola number: %d of'
    % len(fg_with),
    '%d catalog finished goods got one from this workbook.' % len(fg_total),
    'MPP/Hunter to confirm whether FG Macola numbers exist somewhere else.',
])
block('MACOLA # ROWS WITH NO CATALOG PART (%d) -- number is real, part is not '
      'in the MES catalog' % len(macola_unmatched),
      ['%-6s %-32s %-20s macola=%s' % ('r%d' % s['row'], s['desc'][:32], s['pn'] or '(blank)', s['macola'])
       for s in macola_unmatched])
block('BOM QTY CONFLICTS (%d) -- workbook disagrees with the loaded catalog; '
      'NOT emitted' % len(bom_conflict),
      ['%-20s <- %-20s workbook=%s catalog=%s   (%s / %s)'
       % (b['parent'], b['child'], b['qty'], b['prior'], b['fg'], b['desc'])
       for b in bom_conflict])
block('BOM LINKS WHOSE FG NAME DID NOT RESOLVE (%d distinct names)'
      % len({m['fg'] for m in bom_parent_miss + bom_both_miss}),
      sorted({'%-28s %s' % (m['fg'], ('family-sheet PN not in catalog: ' + ', '.join(m['dangling'])) if m.get('dangling') else 'no name match anywhere')
              for m in bom_parent_miss + bom_both_miss}))
block('BOM LINKS WHOSE COMPONENT DID NOT RESOLVE (%d)' % len(bom_child_miss),
      ['%-28s child=%-20s %s' % (m['fg'], m['anchor']['pn'], m['anchor']['desc'][:30])
       for m in bom_child_miss])
A('   ============================================================ */')

sql = '\n'.join(L) + '\n'

# ------------------------------------------------------------- ASCII assert --
non_ascii = [(i + 1, ln) for i, ln in enumerate(sql.split('\n'))
             if any(ord(ch) > 127 for ch in ln)]
if BAD or non_ascii:
    print('ABORT: non-ASCII would be written.')
    for s in BAD[:20]:
        print('  literal:', repr(s))
    for n, ln in non_ascii[:20]:
        print('  line %d: %r' % (n, ln))
    sys.exit(1)

with open(OUT, 'w', encoding='ascii', newline='\n') as f:
    f.write(sql)

# ------------------------------------------------------- reconciliation CSV --
with open(CSV_OUT, 'w', encoding='ascii', newline='') as f:
    w = csv.writer(f)
    w.writerow(['Kind', 'Sheet', 'Row', 'Description', 'HondaPartNumber', 'Macola',
                'FGAssemblyName', 'Qty', 'ResolvedParent', 'ResolvedChild',
                'ResolvedVia', 'Outcome', 'Note'])
    for pn, mac, sheet, row, note in macola_updates:
        w.writerow(['Macola', sheet, row, '', pn, mac, '', '', '', '',
                    'part-number', 'MATCHED', note])
    for a in alloys:
        w.writerow(['Macola', 'ALUMINUM', a['row'], a['desc'], a['pn'], a['macola'],
                    a['fg'], '', '', '', '', 'HELD-BACK',
                    'Raw Material Tracking is FUTURE (Scope Matrix row 21); SQL '
                    'emitted commented out'])
    for s in macola_unmatched:
        w.writerow(['Macola', 'SUPPLY PARTS', s['row'], s['desc'], s['pn'], s['macola'],
                    '', s['qty'], '', '', '', 'UNMATCHED',
                    'Honda Part# has no catalog counterpart'])
    for b in bom_emit:
        w.writerow(['Bom', 'SUPPLY PARTS', b['row'], b['desc'], '', '', b['fg'], b['qty'],
                    b['parent'], b['child'], b['via'], 'EMITTED', 'new BOM line'])
    for b in bom_ok:
        if b['state'] == 'agrees-with-catalog':
            w.writerow(['Bom', 'SUPPLY PARTS', b['row'], b['desc'], '', '', b['fg'], b['qty'],
                        b['parent'], b['child'], b['via'], 'ALREADY-PRESENT',
                        'catalog qty %s' % b['prior']])
    for b in bom_conflict:
        w.writerow(['Bom', 'SUPPLY PARTS', b['row'], b['desc'], '', '', b['fg'], b['qty'],
                    b['parent'], b['child'], b['via'], 'QTY-CONFLICT',
                    'catalog qty %s' % b['prior']])
    for m in bom_parent_miss:
        w.writerow(['Bom', 'SUPPLY PARTS', m['row'], m['anchor']['desc'], m['anchor']['pn'],
                    m['anchor']['macola'], m['fg'], m['qty'], '', '', m['via'],
                    'FG-UNRESOLVED',
                    'family-sheet PN not in catalog: ' + ', '.join(m['dangling']) if m['dangling'] else 'no name match'])
    for m in bom_child_miss:
        w.writerow(['Bom', 'SUPPLY PARTS', m['row'], m['anchor']['desc'], m['anchor']['pn'],
                    m['anchor']['macola'], m['fg'], m['qty'], '', '', '',
                    'COMPONENT-UNRESOLVED', 'Honda Part# has no catalog counterpart'])
    for m in bom_both_miss:
        w.writerow(['Bom', 'SUPPLY PARTS', m['row'], m['anchor']['desc'], m['anchor']['pn'],
                    m['anchor']['macola'], m['fg'], m['qty'], '', '', '',
                    'BOTH-UNRESOLVED',
                    'family-sheet PN not in catalog: ' + ', '.join(m['dangling']) if m['dangling'] else 'no name match'])
    for pn, desc, itype in sorted(no_macola):
        w.writerow(['CatalogNoMacola', '', '', desc, pn, '', '', '', '', '', '',
                    'NO-MACOLA', itype])

# ------------------------------------------------------------------- report --
print('WORKBOOK STRUCTURE')
print('  ALUMINUM      header row 1, %d alloy rows (blank-line separated), MACOLA # at C1'
      % len(alloys))
print('  SUPPLY PARTS  MACOLA # at C1; header band REPEATS at rows %s'
      % ', '.join(str(h) for h in header_rows))
print('                %d anchor rows (carry a MACOLA #), %d component->FG links'
      % (len(supply), len(links)))
print('                %d continuation rows carry their OWN qty (per-FG override)'
      % sum(1 for l in links if l['continuation'] and l['qty_own']))
print('  family sheets %d header bands across %d sheets, %d PART NAME rows'
      % (len(fam_blocks), len(FAMILY_SHEETS), len(fam_pairs)))
for s, i, label in fam_blocks:
    print('                  %-16s row %-3d %s' % (s, i, label))
print()
print('MACOLA        %d values -> %d landed, %d held back (FUTURE scope), %d unmatched'
      % (len(alloys) + len(supply), len(macola_updates), len(alloys), len(macola_unmatched)))
print('              catalog items left with NO Macola: %d' % len(no_macola))
print('              finished goods: %d total, %d given a Macola number'
      % (len(fg_total), len(fg_with)))
print('BOM           %d links -> %d resolved (%d new, %d already present), '
      '%d qty-conflicts' % (len(links), len(bom_ok), len(bom_emit),
                            len(bom_ok) - len(bom_emit), len(bom_conflict)))
print('              unresolved: %d FG-parent, %d component-child, %d both'
      % (len(bom_parent_miss), len(bom_child_miss), len(bom_both_miss)))
print()
print('SCOPE         ALUMINUM (%d alloys) held back -- Scope Matrix row 21, Raw'
      ' Material Tracking = FUTURE. SQL emitted commented out.' % len(alloys))
print('              SERVICE / PASS THRU / NEW MODEL sheets are read for the '
      'PART NAME -> CUSTOMER PART # lookup only;')
print('              nothing is seeded on their behalf (no MACOLA # column, and'
      ' no BOM parent resolved to one).')
if PN_COLLISIONS:
    print()
    print('WARNING       part-number normalization collides %d catalog part(s):'
          % len(PN_COLLISIONS))
    for k, v in sorted(PN_COLLISIONS.items()):
        print('                %s -> %s' % (k, v))
print()
print('wrote %s (%d lines, ASCII verified)' % (OUT, sql.count('\n')))
print('wrote %s' % CSV_OUT)

#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
build_seed_layout_proposal.py
=============================
Transforms the customer-returned part workbook (reference/MPP_Part_Data_Final.xlsx)
into a REVIEW workbook (reference/MPP_Seed_Layout_Proposal.xlsx) laying out exactly
what would be seeded into the Item Master, ahead of writing any SQL.

Decisions baked in (approved 2026-08-12):
  * ItemType   : FinishedGood = MPP's FG rows; SubAssembly = synthesized machined
                 casts; Component = everything else (no PassThrough).
  * SubAssembly: synthesized ONLY where the assembly line has a NON-DEPRECATED
                 Machining OUT terminal. Named "<casting>-M".
  * Die cast   : all castings eligible at DC1 ("die cast one, for now").
  * Trim       : MA1 lines -> TRIM1, MA2 lines -> TRIM2.
  * Ambiguity  : parts that plausibly run on >1 line get an eligibility row at EACH.
  * Off-Site   : loaded as Items + ContainerConfig, but NO route and NO eligibility.
  * Routes     : shape follows sql/seeds/029_seed_item_routes.sql (main), i.e.
                 casting -> DieCast/TrimIn/TrimOut/MachiningIn/(MachiningOut|AssemblyOut)
                 SubAssembly -> MachiningOut ; FinishedGood -> AssemblyOut.

Locations are read from sql/seeds/_site_locations.tsv and passed through the SAME
CODEMAP / skip rules as sql/seeds/gen_locations_mpp.js, so the codes emitted here
match the seeded plant tree exactly.
"""
import os
import re
import sys
import warnings
import collections

warnings.filterwarnings('ignore')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SRC = os.path.join(ROOT, 'reference', 'MPP_Part_Data_Final.xlsx')
TSV = os.path.join(ROOT, 'sql', 'seeds', '_site_locations.tsv')
OUT = os.path.join(ROOT, 'reference', 'MPP_Seed_Layout_Proposal.xlsx')
# Optional output override -- lets the workbook regenerate while the canonical copy
# is open in Excel (openpyxl cannot write a file Excel holds a lock on).
if len(sys.argv) > 1:
    OUT = sys.argv[1]

# ------------------------------------------------------------------
# 1. Location model - mirror gen_locations_mpp.js
# ------------------------------------------------------------------
CODEMAP = {
    'MA1-FP6NA-AFIN': 'MA1-FP6NA-AOUT', 'MA1-FPRPY-AFIN': 'MA1-FPRPY-AOUT',
    'MA2-6F9TC-MOUT': 'MA2-6F9TC-AOUT', 'MA2-COS-MOUT': 'MA2-COS-AOUT',
    'MA2-5PA-MIN1': 'MA2-5PA-MIN', 'MA2-5PA-MIN2': 'MA2-5PA-MOUT',
    'MA2-5PA-MIN3': 'MA2-5PA-AIN', '5PA - AO': 'MA2-5PA-AOUT',
    'MA2-RPY6B2-AFIN': 'MA2-RPY6B2-AOUT',
    'MA2-RPYCAM1-MIN-A': 'MA2-RPYCAM1-MIN-CH', 'MA2-RPYCAM1-MIN-B': 'MA2-RPYCAM1-MIN-RS',
    'MA2-RPYCAM1-AIN': 'MA2-RPYCAM1-MOUT-CH', 'MA2-RPYCAM1-MOUT-B': 'MA2-RPYCAM1-MOUT-RS',
    'MA2-RPYCAM1-MOUT-A': 'MA2-RPYCAM1-MIO-RS5',
    'MA2-RPYCAM1-AOUT1': 'MA2-RPYCAM1-AIN', 'MA2-RPYCAM1-AOUT2': 'MA2-RPYCAM1-AOUT1',
    'MA2-RPYCAM1-AOUT3': 'MA2-RPYCAM1-AOUT2',
    'MA2-RPYCAM2-MOUT-A': 'MA2-RPYCAM2-MIN-CH', 'MA2-RPYCAM2-MIN-B': 'MA2-RPYCAM2-MIN-RS',
    'MA2-RPYCAM2-MOUT-B': 'MA2-RPYCAM2-MOUT-CH', 'RS-MO': 'MA2-RPYCAM2-MOUT-RS',
    'MA2-RPYCAM2-MIN-A': 'MA2-RPYCAM2-MIO-RS5',
}
newcode = lambda c: CODEMAP.get(c, c)
clean = lambda s: re.sub(r'\s+', ' ', s).strip()


def role_of(name, code):
    n = clean(name).lower()
    if re.search(r'in\s*-\s*out', n):
        return 'COMBINED'
    if 'serial' in n:
        return 'ASER'
    if re.search(r'machining\s+out', n):
        return 'MOUT'
    if re.search(r'machining\s+in', n):
        return 'MIN'
    if re.search(r'assembly\s+out', n):
        return 'AOUT'
    if re.search(r'assembly\s+in', n):
        return 'AIN'
    return 'OTHER'


LINES = {}       # lineCode -> {'name':, 'area':, 'terminals': {role: [codes]}}
LINE_NAME = {}
_seen = set()
raw = []
with open(TSV, encoding='utf-8-sig') as fh:
    for ln in fh:
        if not ln.strip():
            continue
        raw.append(ln.rstrip('\n').split('\t'))

kept, deprecated_terms = [], []
for r in raw:
    _id, code, name, ddef, parent, sort, dep = (r + [''] * 7)[:7]
    if not code.strip() or ddef == 'Printer':
        continue
    if dep == '1':
        deprecated_terms.append((newcode(code), clean(name), ddef))
        continue
    if code in _seen:
        continue
    _seen.add(code)
    kept.append((newcode(code), clean(name), ddef, newcode(parent) if parent else None))

DEPRECATED = set(c for c, _n, _d in deprecated_terms)

for code, name, ddef, parent in kept:
    if ddef == 'ProductionLine':
        LINES[code] = {'name': name, 'parent': parent, 'terminals': collections.defaultdict(list)}
        LINE_NAME[code] = name
for code, name, ddef, parent in kept:
    if ddef in ('Terminal', 'InspectionStation') and parent in LINES:
        role = 'INSPECT' if ddef == 'InspectionStation' else role_of(name, code)
        LINES[parent]['terminals'][role].append(code)


def area_of(line):
    p = LINES[line]['parent']
    return p if p in ('MA1', 'MA2') else (p or '')


# a line "mints a SubAssembly" iff it carries a non-deprecated Machining OUT terminal
MOUT_LINES = set(lc for lc, v in LINES.items() if v['terminals'].get('MOUT'))
AIN_LINES = set(lc for lc, v in LINES.items() if v['terminals'].get('AIN'))

# ------------------------------------------------------------------
# 2. Source workbook
# ------------------------------------------------------------------
wb = openpyxl.load_workbook(SRC, data_only=True)


def sheet(name):
    ws, out = wb[name], []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i <= 1:
            continue
        c = ['' if x is None else str(x).strip() for x in r]
        if any(c):
            out.append(c)
    return out


P_ROWS, K_ROWS, B_ROWS = sheet('Parts'), sheet('Packaging'), sheet('BOM')
ISSUES = []


def issue(sev, tab, key, problem, resolution):
    ISSUES.append([sev, tab, key, problem, resolution])


# ---- part-number reconciliation (BOM / Packaging refs -> Parts tab) ----
ALIAS = {
    '1223A-59B -A0002': '1223A-59B-A000',
    '1223A-RPY -G000': '1223A-RPY -A000',
    '1223A-5BA -A001': '1223A-5BA -A000',
    '1223A-5BA -A002': '1223A-5BA -A000',
    '12265-6B2 -A000': '1223A-6B2 -A000',
    '11200-6FB -A000': '11200-6FBA -A000',   # BOM parent -> the FG on Parts
    '11200-6FBA-A000': '11200-6FB-A000',     # BOM child  -> the raw casting on Parts
    '12242-59b-0000': '12242-59B-0000',
}
DROP_PACKAGING = {'5G0-FG', '1223A-59B -A0002'}   # Blue Ridge dev-seed numbers, not MPP parts
norm = lambda pn: ALIAS.get(pn, pn)

for bad, good in ALIAS.items():
    issue('HIGH', 'BOM/Packaging', bad,
          'Part number does not exist on the Parts tab.',
          'Mapped to %r. MPP must confirm.' % good)

# ------------------------------------------------------------------
# 3. Items - dedupe, classify
# ------------------------------------------------------------------
# Parts columns: 0 PN, 1 Desc, 2 Type, 3 Program, 4 Customer, 5 Uom, 6 UnitWeight,
#                7 WeightUom, 8 PartsPerBasket, 9 DefaultSubLotQty, 10 Macola,
#                11 SourceOfTruth, 12 Active, 13 Notes
ITEMS = collections.OrderedDict()

# Two part numbers each cover TWO different parts -- an AEP and an ISP variant with
# different basket sizes and different packing rules. UQ_Item_PartNumber permits one
# row, so they are split by suffixing the variant marker carried in the description.
# Approved 2026-08-13; the suffix is stripped again on the way to AIM.
VARIANT_PARTS = {'19320-6A0 -A510', '19410-6A0 -A000'}
VARIANT_RE = re.compile(r'\((AEP|ISP)\)', re.I)


def variant_pn(pn, desc):
    if pn in VARIANT_PARTS:
        m = VARIANT_RE.search(desc or '')
        if m:
            return '%s-%s' % (pn, m.group(1).upper())
    return pn


for r in P_ROWS:
    pn = variant_pn(r[0], r[1])
    if not pn:
        continue
    if pn not in ITEMS:
        ITEMS[pn] = {
            'pn': pn, 'desc': r[1], 'cust_type': r[2], 'programs': [], 'uom': r[5] or 'EA',
            'unit_weight': r[6], 'weight_uom': r[7], 'basket': r[8], 'sublot': r[9],
            'macola': r[10], 'src': r[11], 'active': r[12], 'notes': r[13],
        }
    it = ITEMS[pn]
    if r[3] and r[3] not in it['programs']:
        it['programs'].append(r[3])
    # keep the richest description / packing values seen
    for fld, idx in (('desc', 1), ('macola', 10), ('basket', 8), ('sublot', 9)):
        if not it[fld] and r[idx]:
            it[fld] = r[idx]
    if r[2] == 'FinishedGood':
        it['cust_type'] = 'FinishedGood'

# duplicate part numbers carrying genuinely different parts -> split into variants
for pn in sorted(VARIANT_PARTS):
    variants = [(r[1], r[8]) for r in P_ROWS if r[0] == pn]
    if len(set(variants)) > 1:
        issue('MED', 'Parts', pn,
              'One part number used for two different parts: %s.' % '; '.join(
                  '%s (basket %s)' % (d, b) for d, b in variants),
              'Split into %s. The suffix is a Blue Ridge convention, NOT an MPP part '
              'number, and MUST be stripped before the value reaches AIM -- '
              'Parts.ufn_AimCustomerPartNumber currently only strips dashes, so it '
              'would emit e.g. "193206A0 A510AEP". MPP should still confirm whether a '
              'real distinct number exists.'
              % ', '.join(sorted(set(variant_pn(pn, d) for d, _b in variants))))
issue('HIGH', 'Parts', '19321-66V -A000 / 19321-66V-A100',
      'Two part numbers with the identical description "66V Thermo Case".',
      'Loaded as two items. MPP to confirm both are live, or retire one.')
issue('MED', 'Parts', 'Program 6AO',
      'Program column has both "6A0" (zero) and "6AO" (letter O).',
      'Treated as one program 6A0.')
issue('MED', 'Parts', '94301-08140',
      'Row has blank Source of Truth and blank Currently Active.',
      'Loaded as active. MPP to confirm.')
issue('MED', 'Parts', '92900-0614-1B',
      'Probable typo - 5PA BOM uses 92900-0614-1B while 6NA/6VJ use 92900-06014-1B '
      '(both described "6x14 Stud Bolt").',
      'Loaded as separate items pending MPP confirmation; likely one part.')
issue('MED', 'Parts', '900009-R70-A000',
      'Probable typo - six leading digits; 6FB BOM uses 90009-R70-A000 for the same '
      'family of plug/bolt.',
      'Loaded as-is pending MPP confirmation.')

# ---- process class: does this part get die cast at MPP? ----
HARDWARE_RE = re.compile(
    r'dowel|bolt|washer|steel ball|o-?ring|stud|plug|joint tube|spring|rubber|'
    r'thermostat|rocker arm assembl|sealing|flange', re.I)
AMBIGUOUS = {'146315GO A000', '146325GO A000', '146335GO A000', '146345GO A000',
             '11221-64AA-A010', '11222-64A-A001'}


BOM_PARENTS = set(norm(r[0]) for r in B_ROWS)


def process_class(it):
    pn, d = it['pn'], it['desc']
    if it['cust_type'] == 'FinishedGood':
        return 'Assembled'
    # A part that is a BOM PARENT is assembled from other parts, whatever MPP typed it.
    # 1223A-6B2 -A000 (the 6B2 Cam Rocker Set) is typed Component but has 20 children;
    # without this it would be classed as a casting and get a die-cast route.
    if pn in BOM_PARENTS:
        return 'Assembled'
    if pn.startswith('P') or pn.startswith('9') or HARDWARE_RE.search(d):
        return 'Purchased'
    return 'Cast'


for it in ITEMS.values():
    it['proc'] = process_class(it)
for pn in sorted(BOM_PARENTS):
    if pn in ITEMS and ITEMS[pn]['cust_type'] != 'FinishedGood':
        issue('HIGH', 'Parts', pn,
              'MPP typed this "%s", but it is a BOM parent with %d components - it is an '
              'assembled part, not a raw one.' % (ITEMS[pn]['cust_type'],
                                                  sum(1 for r in B_ROWS if norm(r[0]) == pn)),
              'Treated as assembled (no die-cast route, no machined identity). MPP should '
              're-type it as FinishedGood.')
for pn in sorted(AMBIGUOUS):
    if pn in ITEMS:
        issue('MED', 'Parts', pn,
              'Cannot tell from the data whether "%s" is die cast at MPP or purchased '
              '(shafts / baffle plates are commonly bought-in steel).' % ITEMS[pn]['desc'],
              'Classified as %s. This drives the die-cast route + DC1/TRIM eligibility, '
              'so MPP should confirm.' % ITEMS[pn]['proc'])

# ------------------------------------------------------------------
# 4. BOM
# ------------------------------------------------------------------
BOM = collections.OrderedDict()
for r in B_ROWS:
    parent, child, qty, note = norm(r[0]), norm(r[1]), r[2], r[3]
    if parent == child:
        issue('HIGH', 'BOM', parent,
              'BOM row lists the part as its own component (self-reference).',
              'Row dropped. MPP to supply the intended component.')
        continue
    if parent not in ITEMS:
        issue('HIGH', 'BOM', parent, 'BOM parent not resolvable to any item.', 'Row dropped.')
        continue
    if child not in ITEMS:
        issue('HIGH', 'BOM', child, 'BOM child not resolvable to any item.', 'Row dropped.')
        continue
    BOM.setdefault(parent, []).append({'child': child, 'qty': qty, 'note': note})

for pn, it in ITEMS.items():
    if it['cust_type'] == 'FinishedGood' and pn not in BOM:
        sev = 'LOW' if it['src'] == 'Off-Site' else 'HIGH'
        issue(sev, 'BOM', pn,
              'Finished good has no BOM rows.',
              'Off-site part - no BOM expected.' if sev == 'LOW'
              else 'Cannot build an assembly BOM. MPP must supply the make-up.')

# note-vs-description mismatches inside the BOM
for parent, lines in BOM.items():
    for l in lines:
        n, d = clean(l['note']).lower(), clean(ITEMS[l['child']]['desc']).lower()
        if n and d and n != d and not (n in d or d in n):
            if re.match(r'^\d+x\d+', n) and re.match(r'^\d+x\d+', d) and n.split()[0] != d.split()[0]:
                issue('MED', 'BOM', '%s -> %s' % (parent, l['child']),
                      'BOM note says "%s" but the part is "%s".' % (l['note'], ITEMS[l['child']]['desc']),
                      'Item description kept. MPP to confirm the correct component.')

# ------------------------------------------------------------------
# 5. Location matching
# ------------------------------------------------------------------
# FG -> line(s). Derived from the authoritative line NAMES in _site_locations.tsv.
FG_LINES = {
    '1223A-59B-A000':   ['MA2-59B'],
    '1223A-RPY -A000':  ['MA2-RPYCAM1', 'MA2-RPYCAM2'],
    '1223A-5BA -A000':  ['MA2-RPYCAM1', 'MA2-RPYCAM2'],
    '1223A-6B2 -A000':  ['MA2-RPY6B2'],
    '1223A-6MA -J000':  ['MA2-6MACH', '6ma-CH-L2'],
    '14650-5GO -A000':  ['MA1-5GOF'],
    '14660-5GO -A000':  ['MA1-5GOR'],
    '12270-6NA -A000':  ['MA1-FP6NA'],
    '12270-6VJ -A000':  ['MA1-FP6NA'],
    '12270-5PA -A000':  ['MA2-5PA'],
    '12270-66V -A000':  ['MA1-FPRPY'],
    '12270-RPY -G001':  ['MA1-FPRPY'],
    '12270-6B2 -A000':  ['MA2-RPY6B2'],
    '17145-6MD -A000':  ['MA1-6MD'],
    '1120A-64A -A000':  ['MA2-64AOP'],
    '1120A-69F -A000':  ['MA2-6MAOP'],
    '1932A-69F -A000':  ['MA2-6F9TC'],
    '12235-6FB -A000':  ['MA2-6FBCHOP'],
    '12431-6FB -A000':  ['MA2-6FBCHOP'],
    '11200-5J6 -A110':  ['MA2-V6OP'],
}
MATCH_WHY = {
    '1223A-59B-A000':  'line "59b Cam holder"',
    '1223A-RPY -A000': 'lines "RPY Line 1/2 Cam Holders" - two identical lines, eligible at both',
    '1223A-5BA -A000': 'no 5BA line; built from RPY components, so both RPY cam-holder lines',
    '1223A-6B2 -A000': 'line "RPY 6b2 line2"',
    '1223A-6MA -J000': 'lines "6MA Cam Holder Line 1" + "6MA Cam Holder Line 2"',
    '14650-5GO -A000': 'line "5G0 Front" (part is RKR SHAFT FR)',
    '14660-5GO -A000': 'line "5G0 Rear" (part is RKR Shaft RR)',
    '12270-6NA -A000': 'line "Fuel Pump (6NA 6VJ)"',
    '12270-6VJ -A000': 'line "Fuel Pump (6NA 6VJ)"',
    '12270-5PA -A000': 'line "5PA Fuel Pump"',
    '12270-66V -A000': 'line "Fuel Pump (RPY 66v)"',
    '12270-RPY -G001': 'line "Fuel Pump (RPY 66v)"',
    '12270-6B2 -A000': 'WEAK - only 6B2 line is "RPY 6b2 line2", which looks like a cam-holder line',
    '17145-6MD -A000': 'line "6MD Manifold Plate"',
    '1120A-64A -A000': 'line "64A Oil Pan"',
    '1120A-69F -A000': 'line "6ma oil pan" - the BOM raw is "6MA Oil Pan Raw" (11200-6MAA-J010)',
    '1932A-69F -A000': 'WEAK - line "6F9-TC"; 6F9 vs 69F looks like transposed digits',
    '12235-6FB -A000': 'line "6FB Small Parts", child line "Assembly Out Oil Passage"',
    '12431-6FB -A000': 'line "6FB Small Parts", terminal "Assembly Out Cam Holders"',
    '11200-5J6 -A110': 'WEAK - only unassigned oil-pan line is "v6 Oil Pan"',
}
for pn in FG_LINES:
    for lc in FG_LINES[pn]:
        if lc not in LINES:
            raise SystemExit('BUG: line code %r not in the location model' % lc)

# propagate FG -> its BOM children, then Program -> FG as a fallback
lines_for = {}
for pn, lcs in FG_LINES.items():
    lines_for.setdefault(pn, set()).update(lcs)
for _ in range(3):   # BOM depth is shallow; 3 passes is ample
    for parent, kids in BOM.items():
        if parent not in lines_for:
            continue
        for k in kids:
            lines_for.setdefault(k['child'], set()).update(lines_for[parent])
prog_lines = collections.defaultdict(set)
for pn, lcs in FG_LINES.items():
    for p in ITEMS[pn]['programs']:
        prog_lines[p].update(lcs)
prog_lines['6AO'] = prog_lines['6A0']
for pn, it in ITEMS.items():
    if pn in lines_for:
        continue
    # The program fallback is for COMPONENTS only. A finished good shares its program
    # code with unrelated finished goods (6FB covers both "6FB Small Parts" and an oil
    # pan), so falling back on it invents a line assignment. An FG must be matched by
    # its own name in FG_LINES, or it belongs on the Unmatched tab.
    if it['cust_type'] == 'FinishedGood':
        continue
    # Candidate finished goods sharing this component's program code.
    cand_fg = [f for f in FG_LINES if set(ITEMS[f]['programs']) & set(it['programs'])]
    if not cand_fg:
        continue
    # A program code can cover unrelated finished goods (RPY covers both the cam-holder
    # set and the fuel pump), so spreading a component across all of them puts a fuel-pump
    # casting on a cam-holder line. When the component's own description names one of the
    # candidates ("RPY Fuel Pump Raw" -> "RPY Fuel Pump"), trust that and only that.
    stem = re.sub(r'\b(raw|casting)\b', '', clean(it['desc']), flags=re.I).strip().lower()
    named = [f for f in cand_fg if stem and clean(ITEMS[f]['desc']).lower() == stem]
    if named:
        cand_fg, why = named, 'description matches finished good %s' % named[0]
    else:
        why = 'via program %s - no BOM parent, so every %s finished good is a candidate' % (
            '/'.join(it['programs']), '/'.join(it['programs']))
        if len(cand_fg) > 1:
            issue('MED', 'BOM', pn,
                  'Component "%s" appears in no BOM, so its line was inferred from program '
                  '%s, which covers %d finished goods.' % (it['desc'], '/'.join(it['programs']),
                                                           len(cand_fg)),
                  'Eligible at all of them: %s. MPP should supply the BOM.'
                  % ', '.join(sorted(set(l for f in cand_fg for l in FG_LINES[f]))))
    cand = set(l for f in cand_fg for l in FG_LINES[f])
    if cand:
        lines_for[pn] = cand
        it['match_note'] = why

OFFSITE = set(pn for pn, it in ITEMS.items() if it['src'] == 'Off-Site')
for pn in OFFSITE:
    lines_for.pop(pn, None)          # off-site: no eligibility, no route

UNMATCHED = [pn for pn in ITEMS if pn not in lines_for]

# ------------------------------------------------------------------
# 6. SubAssembly synthesis
# ------------------------------------------------------------------
SUBS = collections.OrderedDict()     # casting pn -> sub pn
for pn, it in ITEMS.items():
    if it['proc'] != 'Cast' or pn not in lines_for:
        continue
    if any(lc in MOUT_LINES for lc in lines_for[pn]):
        SUBS[pn] = pn + '-M'

for cast, sub in SUBS.items():
    src = ITEMS[cast]
    ITEMS[sub] = {
        'pn': sub, 'desc': (src['desc'] + ' (machined)')[:500], 'cust_type': '',
        'programs': list(src['programs']), 'uom': src['uom'], 'unit_weight': '',
        'weight_uom': '', 'basket': src['basket'], 'sublot': src['sublot'],
        'macola': '', 'src': 'Blue Ridge (synthesized)', 'active': 'Yes',
        'notes': 'Machined identity minted by Machining OUT from %s.' % cast,
        'proc': 'Machined', 'synth': True,
    }
    lines_for[sub] = set(lc for lc in lines_for[cast] if lc in MOUT_LINES)


def item_type(pn):
    it = ITEMS[pn]
    if it.get('synth'):
        return 'SubAssembly'
    return 'FinishedGood' if it['cust_type'] == 'FinishedGood' else 'Component'


# ------------------------------------------------------------------
# 7. Routes  (shape per sql/seeds/029_seed_item_routes.sql on main)
# ------------------------------------------------------------------
ROUTES = []
for pn, it in ITEMS.items():
    t = item_type(pn)
    lcs = sorted(lines_for.get(pn, ()))
    if t == 'FinishedGood':
        if pn in OFFSITE:
            continue
        ROUTES.append((pn, 1, 'AssemblyOut', 'Assembly out', ''))
    elif t == 'SubAssembly':
        ROUTES.append((pn, 1, 'MachiningOut', 'Machining out', ''))
    elif it['proc'] == 'Cast' and lcs:
        area = set(area_of(lc) for lc in lcs)
        trim = 'TRIM1' if area == {'MA1'} else ('TRIM2' if area == {'MA2'} else 'TRIM?')
        steps = [('DieCast', 'Die cast'), ('TrimIn', 'Trim in at %s' % trim),
                 ('TrimOut', 'Trim out at %s' % trim), ('MachiningIn', 'Machining in')]
        if pn in SUBS:
            steps.append(('MachiningOut', 'Machining out (mints %s)' % SUBS[pn]))
            mint = SUBS[pn]
        else:
            steps.append(('AssemblyOut', 'Assembly out'))
            mint = ''
        for i, (role, d) in enumerate(steps, 1):
            ROUTES.append((pn, i, role, d, mint if i == len(steps) else ''))

# ------------------------------------------------------------------
# 8. Eligibility
# ------------------------------------------------------------------
ELIG = []
for pn, it in ITEMS.items():
    t, lcs = item_type(pn), sorted(lines_for.get(pn, ()))
    if not lcs:
        continue
    if it['proc'] == 'Cast':
        areas = sorted(set(area_of(lc) for lc in lcs))
        ELIG.append((pn, 'DC1', 'Die Cast 1', 'ProductionArea', 0,
                     'All castings die cast at DC1 (interim decision).'))
        for a in areas:
            trim = 'TRIM1' if a == 'MA1' else ('TRIM2' if a == 'MA2' else None)
            if trim:
                ELIG.append((pn, trim, 'Trim Shop %s' % trim[-1], 'ProductionArea', 0,
                             'Assembly line is under %s.' % a))
    for lc in lcs:
        cp = 1 if (t == 'Component' and it['proc'] in ('Purchased', 'Cast')) else 0
        ELIG.append((pn, lc, LINE_NAME[lc], 'ProductionLine', cp,
                     MATCH_WHY.get(pn) or it.get('match_note') or 'via BOM parent'))

# ------------------------------------------------------------------
# 9. Packaging
# ------------------------------------------------------------------
CLOSURE = {'camera': 'ByVision', 'scale': 'ByWeight', 'bycount': 'ByCount',
           'byweight': 'ByWeight', 'byvision': 'ByVision'}
PACK, seen_pack = [], set()
for r in K_ROWS:
    pn = norm(r[0])
    if r[0] in DROP_PACKAGING or pn in DROP_PACKAGING:
        issue('MED', 'Packaging', r[0],
              'Blue Ridge dev-seed part number appears in the customer workbook.',
              'Row dropped - not an MPP part.')
        continue
    # A split part number carries one packaging row per variant. Route each to its
    # variant by basket size: MPP's Parts "Parts Per Basket" equals the Packaging
    # "Parts Per Tray" for both of these (AEP 5 -> 16x5 / ISP 6 -> 16x6;
    # AEP 60 -> 1x60 / ISP 15 -> 12x15), so the match is determined, not guessed.
    if pn in VARIANT_PARTS:
        cands = [k for k in ITEMS if k.startswith(pn + '-')
                 and str(ITEMS[k]['basket']).strip() == str(r[2]).strip()]
        if len(cands) == 1:
            pn = cands[0]
        else:
            issue('HIGH', 'Packaging', r[0],
                  'Packaging row (%s trays x %s parts) does not match exactly one '
                  'AEP/ISP variant by basket size.' % (r[1], r[2]),
                  'Row dropped - resolve the variant split first.')
            continue
    if pn not in ITEMS:
        issue('HIGH', 'Packaging', r[0], 'Packaging row for an unknown part.', 'Row dropped.')
        continue
    if pn in seen_pack:
        issue('HIGH', 'Packaging', pn,
              'Second packaging row for the same part (%s trays x %s parts, %s).'
              % (r[1], r[2], r[4]),
              'UQ_ContainerConfig_ActiveItemId permits one active config. Row dropped; '
              'MPP must split the part number or pick one rule.')
        continue
    seen_pack.add(pn)
    closure = CLOSURE.get(clean(r[4]).lower(), '')
    if not closure and r[4]:
        issue('MED', 'Packaging', pn, 'Unrecognised closure method %r.' % r[4], 'Left blank.')
    tw = r[5]
    if closure == 'ByWeight' and not tw:
        issue('MED', 'Packaging', pn,
              'Closure is "Scale" (ByWeight) but no Target Weight was given.',
              'TargetWeight left NULL - the scale close cannot validate without it.')
    PACK.append([pn, ITEMS[pn]['desc'], r[1], r[2], 1 if clean(r[3]).lower() == 'yes' else 0,
                 closure, tw, r[6], r[7], r[4]])

for pn, it in ITEMS.items():
    if item_type(pn) == 'FinishedGood' and pn not in seen_pack:
        issue('HIGH', 'Packaging', pn, 'Finished good has no packaging rule.',
              'No ContainerConfig can be created; containers cannot close.')

# unused lines
used = set()
for s in lines_for.values():
    used.update(s)
for lc in sorted(LINES):
    if lc not in used:
        issue('MED', 'Locations', lc,
              'Configured production line "%s" has no part assigned to it.' % LINE_NAME[lc],
              'Either MPP omitted its parts, or the line is idle.')

# ------------------------------------------------------------------
# 10. Emit
# ------------------------------------------------------------------
HDR_FILL = PatternFill('solid', fgColor='1F3864')
HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
SUB_FILL = PatternFill('solid', fgColor='FFF2CC')
FLAG_FILL = PatternFill('solid', fgColor='FCE4E4')
OK_FILL = PatternFill('solid', fgColor='E2EFDA')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

out = openpyxl.Workbook()
out.remove(out.active)


def add(title, headers, rows, widths=None, freeze='A2', fills=None):
    ws = out.create_sheet(title)
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BORDER
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths or [], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = BORDER
            c.alignment = Alignment(vertical='top', wrap_text=False)
    if fills:
        for ridx, row in enumerate(ws.iter_rows(min_row=2), 2):
            f = fills(rows[ridx - 2])
            if f:
                for c in row:
                    c.fill = f
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    return ws


# --- Items ---
irows = []
for pn, it in ITEMS.items():
    t = item_type(pn)
    lcs = sorted(lines_for.get(pn, ()))
    status = ('SYNTHESIZED' if it.get('synth') else
              'OFF-SITE' if pn in OFFSITE else
              'UNMATCHED' if not lcs else 'OK')
    irows.append([
        pn, it['desc'], t, it['cust_type'] or '-', it['proc'],
        '/'.join(it['programs']), it['uom'], it['unit_weight'], it['weight_uom'],
        it['basket'], it['sublot'], it['macola'], it['src'], it['active'],
        ', '.join(lcs) or '(none)', status, it['notes'],
    ])
add('Items',
    ['Part Number', 'Description', 'Proposed ItemType', 'MPP ItemType', 'Process Class',
     'Program', 'Uom', 'Unit Weight', 'Weight Uom', 'MaxLotSize (Parts/Basket)',
     'DefaultSubLotQty', 'Macola #', 'Source of Truth', 'Active', 'Assigned Line(s)',
     'Status', 'Notes'],
    irows,
    [22, 34, 17, 14, 13, 10, 6, 10, 10, 20, 16, 12, 18, 8, 30, 13, 46],
    fills=lambda r: (SUB_FILL if r[15] == 'SYNTHESIZED'
                     else FLAG_FILL if r[15] in ('UNMATCHED', 'OFF-SITE') else None))

# --- Eligibility ---
add('Eligibility',
    ['Part Number', 'Location Code', 'Location Name', 'Tier', 'IsConsumptionPoint', 'Why'],
    [list(e) for e in ELIG], [22, 24, 30, 16, 20, 72])

# --- Routes ---
rrows = []
for pn, seq, role, desc, mint in ROUTES:
    rrows.append([pn, '%s Route v1' % pn, seq, role, desc, mint])
add('Routes',
    ['Part Number', 'Route Name', 'Seq', 'OperationType Role', 'Step Description',
     'Mints Part Number'],
    rrows, [22, 34, 6, 20, 40, 24])

# --- BOMs ---
brows = []
for parent, kids in BOM.items():
    for i, k in enumerate(kids, 1):
        child = k['child']
        # a casting consumed on a MOUT line is consumed as its MACHINED identity
        eff = SUBS.get(child, child)
        brows.append([parent, ITEMS[parent]['desc'], eff, ITEMS[eff]['desc'], k['qty'], i,
                      'customer' if eff == child else 'retargeted to machined identity',
                      k['note']])
add('BOMs',
    ['Parent Part Number', 'Parent Description', 'Child Part Number', 'Child Description',
     'Qty Per', 'Sort', 'Origin', 'MPP Note'],
    brows, [22, 32, 22, 32, 9, 6, 30, 30],
    fills=lambda r: SUB_FILL if r[6] != 'customer' else None)

# --- Packaging ---
add('Packaging',
    ['Part Number', 'Description', 'TraysPerContainer', 'PartsPerTray', 'IsSerialized',
     'ClosureMethod', 'TargetWeight', 'DunnageCode', 'CustomerCode', 'MPP raw closure'],
    PACK, [22, 32, 18, 14, 13, 15, 13, 14, 14, 16],
    fills=lambda r: FLAG_FILL if (r[5] == 'ByWeight' and not r[6]) else None)

# --- Unmatched ---
urows = []
for pn in UNMATCHED:
    it = ITEMS[pn]
    urows.append([pn, it['desc'], item_type(pn), '/'.join(it['programs']), it['src'],
                  'Off-site part - manufactured away from Madison; no MPP line applies.'
                  if pn in OFFSITE else
                  'No line in the location model matches this part by name, and it is '
                  'not a BOM child of any matched finished good.'])
add('Unmatched',
    ['Part Number', 'Description', 'Proposed ItemType', 'Program', 'Source of Truth', 'Why unmatched'],
    urows, [22, 34, 17, 10, 16, 76], fills=lambda r: FLAG_FILL)

# --- DataIssues ---
sev_order = {'HIGH': 0, 'MED': 1, 'LOW': 2}
ISSUES.sort(key=lambda x: (sev_order.get(x[0], 9), x[1], x[2]))
add('DataIssues', ['Severity', 'Tab', 'Key', 'Problem', 'Proposed resolution'],
    ISSUES, [10, 16, 28, 74, 74],
    fills=lambda r: FLAG_FILL if r[0] == 'HIGH' else (SUB_FILL if r[0] == 'MED' else None))

# --- Summary (first tab) ---
n_fg = sum(1 for pn in ITEMS if item_type(pn) == 'FinishedGood')
n_sa = sum(1 for pn in ITEMS if item_type(pn) == 'SubAssembly')
n_co = sum(1 for pn in ITEMS if item_type(pn) == 'Component')
srows = [
    ['SOURCE', 'reference/MPP_Part_Data_Final.xlsx', ''],
    ['Rows returned by MPP', '%d Parts / %d Packaging / %d BOM' % (len(P_ROWS), len(K_ROWS), len(B_ROWS)), ''],
    ['', '', ''],
    ['PROPOSED ITEM MASTER', '%d items' % len(ITEMS), ''],
    ['  FinishedGood', n_fg, 'Rows MPP marked FinishedGood'],
    ['  SubAssembly', n_sa, 'Synthesized machined identities (highlighted amber)'],
    ['  Component', n_co, 'Castings + purchased hardware'],
    ['', '', ''],
    ['Eligibility rows', len(ELIG), 'Item x Location'],
    ['Route steps', len(ROUTES), 'Across %d routed items' % len(set(r[0] for r in ROUTES))],
    ['BOM lines', len(brows), 'Across %d parents' % len(BOM)],
    ['Container configs', len(PACK), 'One per finished good'],
    ['', '', ''],
    ['Unmatched parts', len(UNMATCHED), 'See Unmatched tab - %d are off-site' % len(OFFSITE)],
    ['Data issues', len(ISSUES),
     '%d HIGH / %d MED / %d LOW' % (sum(1 for i in ISSUES if i[0] == 'HIGH'),
                                    sum(1 for i in ISSUES if i[0] == 'MED'),
                                    sum(1 for i in ISSUES if i[0] == 'LOW'))],
    ['', '', ''],
    ['RULES APPLIED', '', ''],
    ['Item types', 'FinishedGood / SubAssembly / Component',
     'Finished parts = FinishedGood; machined casts = SubAssembly; everything else = Component.'],
    ['SubAssembly synthesis', 'Active MOUT lines only',
     'Lines with a non-deprecated Machining OUT terminal: ' + ', '.join(sorted(MOUT_LINES))],
    ['Die cast eligibility', 'DC1', 'All castings, per the interim "die cast one, for now" decision.'],
    ['Trim eligibility', 'MA1 -> TRIM1, MA2 -> TRIM2', 'By the parent area of the assigned line.'],
    ['Ambiguous parts', 'Eligible at every candidate line', 'No single-line guess is forced.'],
    ['Off-site parts', 'Item + ContainerConfig only', 'No route, no eligibility.'],
    ['Route shape', 'per sql/seeds/029_seed_item_routes.sql',
     'Casting: DieCast/TrimIn/TrimOut/MachiningIn/(MachiningOut|AssemblyOut). '
     'SubAssembly: MachiningOut. FinishedGood: AssemblyOut.'],
    ['Location source', 'sql/seeds/_site_locations.tsv',
     'Codes passed through the same CODEMAP + skip rules as gen_locations_mpp.js.'],
]
ws = add('Summary', ['Item', 'Value', 'Detail'], srows, [30, 46, 96], freeze='A2')
out._sheets.insert(0, out._sheets.pop(out._sheets.index(ws)))

out.save(OUT)
print('Wrote %s' % OUT)
print('  items=%d (FG=%d SA=%d C=%d)  elig=%d  routes=%d  bom=%d  pack=%d  unmatched=%d  issues=%d'
      % (len(ITEMS), n_fg, n_sa, n_co, len(ELIG), len(ROUTES), len(brows), len(PACK),
         len(UNMATCHED), len(ISSUES)))
